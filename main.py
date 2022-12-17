import pandas as pd
from openpyxl import *
from openpyxl.utils import get_column_letter
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import *
from nomen import *
from openpyxl.styles import NamedStyle
import jinja2
import re
import copy
import numpy as np
from pandas.io.excel import ExcelWriter
from tkinter import *


def format(lst, let, num, height, width, size, bold):
    lst.row_dimensions[num].height = height
    lst.column_dimensions[str(let)].width = width
    lst[let + str(num)].alignment = Alignment(horizontal='center')
    lst[let + str(num)].font = Font(size=size, bold=bold)
    lst[let + str(num)].alignment = Alignment(wrap_text=True)


ost = load_workbook('СХТ.xlsx')
subp = load_workbook('Субпродукты.xlsx')
subi = subp['TDSheet']
sub = subp.create_sheet('пох', 0)
asub = 1
for i in range(12, subi.max_row):
    nom = subi['B' + str(i)].value
    kol = subi['G' + str(i)].value
    if nom and kol and not re.search('ерев', nom) and not re.search('пузыр', nom):
        sub['A' + str(asub)] = str(nom)
        sub['B' + str(asub)] = float(str(kol))
        asub += 1

dor = ost.create_sheet('Доработка')
dor = ost['Доработка']
ost_pt = ost.create_sheet('Остатки полутуш', 0)
ost_mb = ost.create_sheet('Остатки по датам молодняк', -1)
ost_vk = ost.create_sheet('Остаток по датам коровы', -1)
livG = ost.create_sheet('Ливер говядина', -1)
livG = ost['Ливер говядина']
livS = ost.create_sheet('Ливер свинина', -1)
livS = ost['Ливер свинина']
maso = ost.create_sheet('Остатки мяса и костей', -1)
maso = ost['Остатки мяса и костей']

memory = ost.create_sheet('Памятка')
memory = ost['Памятка']
pam = []
date = datetime.datetime.now().strftime('%d.%m.%Y')
birn = ost['TDSheet']
name = f'Остатки на {date}'
sl = 4
gl = 4
ddor = 1
ms = 4


def click():
    maso['B' + str(ms + 1)] = int(count.get())
    window.destroy()


for i in range(2, sub.max_row):
    nom = sub['A' + str(i)].value
    pp = sub['A' + str(i + 1)].value
    ppk = sub['B' + str(i + 1)].value
    kol = sub['B' + str(i)].value
    if kol < 0:
        pam.append('На складе субпродуктов отрицательные остатки, нужно исправить!!!')
    if nom in livGG:
        livG['A' + str(gl)] = nom
        livG['B' + str(gl)] = kol
        gl += 1
    elif 'Пром.' in pp and nom in livGG:
        livG['A' + str(gl)] = nom + '  ПП'
        livG['B' + str(gl)] = ppk
        livG['B' + str(gl - 1)] = kol - ppk
        gl += 1

    elif nom in livSS:
        livS['A' + str(sl)] = nom
        livS['B' + str(sl)] = kol
        sl += 1
    elif 'Пром.' in pp and nom in livSS:
        livS['A' + str(sl)] = nom + '  ПП'
        livS['B' + str(sl)] = ppk
        livS['B' + str(sl - 1)] = kol - ppk
        sl += 1
    elif nom in dorob:
        if kol > 20:
            pam.append(
                f'{nom} Доработки больше 20 кг, зайди на лист доработка и проанализируй. Сообщи об этом менеджеру или зав.складом , если необходимо. ')  # memory1
        dor['A' + str(ddor)] = nom
        dor['B' + str(ddor)] = kol
        ddor += 1
    elif nom in zam:
        maso['A' + str(ms)] = nom
        maso['C' + str(ms)] = kol
        maso['A' + str(ms)].alignment = Alignment(wrap_text=True)
        ms += 1
    elif not nom in pobocka and not re.search('ром. переработка', nom):
        if re.search('Гов. 2 с. колбасная', nom):
            maso['A' + str(ms)] = nom
            maso['B' + str(ms)] = kol
            maso['A' + str(ms)].alignment = Alignment(wrap_text=True)

            window = Tk()
            window.geometry('600x600')
            lbl = Label(window, text=f'Укажи рульку {nom}?', font=("Arial Bold", 15))
            lbl.pack()

            count = Entry(window)
            count.pack()
            btn = Button(window, text="ОК!", command=click).pack()

            window.mainloop()
            maso['A' + str(ms + 1)] = f'{nom}  Рулька'
            maso['B' + str(ms)] = maso['B' + str(ms)].value - maso['B' + str(ms + 1)].value
            ms += 2
        if re.search('Св. жирная колб', nom) or re.search('Св. п/ж колб', nom):
            maso['A' + str(ms)] = nom
            maso['B' + str(ms)] = kol
            maso['A' + str(ms)].alignment = Alignment(wrap_text=True)

            window = Tk()
            window.geometry('600x600')
            lbl = Label(window, text=f'Кол-во кров и мелкой {nom}?', font=("Arial Bold", 15))
            lbl.pack()

            count = Entry(window)
            count.pack()
            btn = Button(window, text="ОК!", command=click).pack()

            window.mainloop()
            maso['A' + str(ms + 1)] = f'{nom}  КРОВ/МЕЛКАЯ'
            maso['B' + str(ms)] = maso['B' + str(ms)].value - maso['B' + str(ms + 1)].value
            ms += 2
        else:
            maso['A' + str(ms)] = nom
            maso['B' + str(ms)] = kol
            maso['A' + str(ms)].alignment = Alignment(wrap_text=True)
            ms += 1
ost.create_sheet('Брыжейка')
ost.create_sheet('Нежил')

dor.column_dimensions['A'].width = 65
livS.column_dimensions['A'].width = 55
livG.column_dimensions['A'].width = 55
maso.column_dimensions['A'].width = 80
for i in [livS, livG, maso]:
    i.merge_cells('A1:A2')
    i['A1'] = 'Наименование'
    i.merge_cells('B1:B2')
    i['B1'] = 'ОХЛ'
    i.merge_cells('C1:C2')
    i['C1'] = 'ЗАМ'
    i.merge_cells('D1:E1')
    i.column_dimensions['D'].width = 15
    i.column_dimensions['E'].width = 15
    i['D2'] = 'Ливерный'
    i['E2'] = 'Склад № 2'
    i['D1'] = 'Заполняет менеджер'
    for row in range(1, i.max_row + 1):
        for column in 'ABCDE':
            thack = Side(border_style='medium', color="0A0000")
            i[column + str(row)].border = Border(top=thack,
                                                 left=thack,
                                                 right=thack,
                                                 bottom=thack)
            i.row_dimensions[row].height = 24
            i[column + str(row)].font = Font(size=9, bold=True)
            i['D2'].font = Font(size=11, italic=True, bold=True, color='871D1D')
            i['E2'].font = Font(size=11, italic=True, bold=True, color='871D1D')
            i['A1'].font = Font(size=14, italic=True, bold=True, color='871D1D')
            i['B1'].font = Font(size=14, italic=True, bold=True, color='871D1D')
            i['C1'].font = Font(size=14, italic=True, bold=True, color='871D1D')
            i['D1'].font = Font(size=11, italic=True, bold=True, color='871D1D')
maso['D2'] = 'Примечание'
maso['E2'] = 'Приоритет'
ost_pt['A1'] = 'Остатки холодильника хранения туш '

format(ost_pt, 'A', 1, 23, 65, 16, True)


def oform(lst, let, num, thickless):
    thack = Side(border_style=thickless, color="0A0000")
    lst[let + str(num)].border = Border(top=thack,
                                        left=thack,
                                        right=thack,
                                        bottom=thack)


oform(ost_pt, 'A', 1, 'thick')
# добавляет 0
for i in range(11, birn.max_row):
    try:
        birn['J' + str(i)].value * 1
    except:
        birn['J' + str(i)].value = 0
ost_pt.merge_cells('A1:B1')

ost_pt['B2'] = 'Кол-во туш'
format(ost_pt, 'B', 2, 18, 26, 14, True)
oform(ost_pt, 'B', 2, 'medium')
ost_pt['D2'] = 'Кол-во туш'
format(ost_pt, 'D', 2, 18, 26, 14, True)
oform(ost_pt, 'D', 2, 'medium')
ost_pt['D1'] = 'Сан.камера'
format(ost_pt, 'D', 1, 18, 37, 16, True)
oform(ost_pt, 'D', 1, 'medium')
for i in range(1, 14):
    ost_pt['D' + str(i)].fill = PatternFill('solid', start_color="F7FF00")
for i in range(3, 14):
    for j in range(1, 5):
        format(ost_pt, get_column_letter(j), i, 18, 36, 11, True)
        oform(ost_pt, get_column_letter(j), i, 'thin')
for i in range(1, 4):
    ost_pt[get_column_letter(i) + str(2)].fill = PatternFill('solid',
                                                             start_color="75F286")
ost_pt['C2'] = 'Дата'
ost_pt['C2'].font = Font(size=16, bold=True)

for i in range(3, 13):
    ost_pt['A' + str(i)] = list(sv.keys())[i - 3]

format(ost_pt, 'A', 13, 23, 65, 16, True)
oform(ost_pt, 'A', 13, 'thick')
ost_pt['A13'] = 'Итого'
for i in [7, 10, 13]:
    for j in range(1, 4):
        ost_pt[get_column_letter(j) + str(i)].fill = PatternFill(
            'solid', start_color='BCBCBC')
for i in range(19, 32):
    for j in range(1, 3):
        format(ost_pt, get_column_letter(j), i, 18, 36, 11, True)
        oform(ost_pt, get_column_letter(j), i, 'thin')

format(ost_pt, 'A', 32, 23, 65, 16, True)
oform(ost_pt, 'A', 32, 'thick')
ost_pt['A32'] = 'Итого'
ost_pt['A32'].fill = PatternFill('solid', start_color='BCBCBC')
ost_pt['B32'].fill = PatternFill('solid', start_color='BCBCBC')
format(ost_pt, 'A', 13, 23, 65, 16, False)

for i in range(19, 32):
    ost_pt['A' + str(i)] = KrsSV[i - 19]

ost_pt['A14'] = 'Свиноматки б/ш'
ost_pt['A15'] = 'Свиноматки в/ш'
ost_pt['A16'] = '2 б/ш Староминка'
ost_pt['A14'].font = Font(size=11)
ost_pt['A15'].font = Font(size=11)
ost_pt['A16'].font = Font(size=11)
for j in range(14, 16):
    format(ost_pt, 'A', i, 18, 36, 11, True)
format(ost_pt, 'A', 32, 23, 65, 16, True)
oform(ost_pt, 'A', 32, 'thick')
format(ost_pt, 'A', 32, 23, 65, 16, True)
ost_pt['A35'] = 'Браки'
ost_pt['B35'] = 'Кол-во п/т / четвертин'
ost_pt['A35'].fill = PatternFill('solid', start_color='7AF38A')
ost_pt['B35'].fill = PatternFill('solid', start_color='7AF38A')
oform(ost_pt, 'A', 35, 'thick')
format(ost_pt, 'A', 35, 23, 65, 16, True)
oform(ost_pt, 'B', 35, 'thick')
format(ost_pt, 'B', 35, 23, 65, 16, True)

a = 36
for i in range(1, birn.max_row):
    if re.search('без баков', str(birn['B' + str(i)].value)) and not re.search(
            'виномат|Староми', str(birn['B' + str(i)].value)):
        ost_pt['A' + str(a)] = birn['B' + str(i)].value
        ost_pt['B' + str(a)] = birn['J' + str(i)].value

        a += 1
for i in range(36, a):
    for j in range(1, 2):
        format(ost_pt, get_column_letter(j), i, 34, 36, 8, False)
        oform(ost_pt, get_column_letter(j), i, 'thin')
ost_pt['A' + str(a)] = 'Итого'
ost_pt['A' + str(a)].fill = PatternFill('solid', start_color='BCBCBC')
ost_pt['B' + str(a)].fill = PatternFill('solid', start_color='BCBCBC')

ost_pt['B' + str(a)] = sum([
    int(ost_pt['B' + str(i)].value) for i in range(36, a)
    if ost_pt['B' + str(i)].value != None
])
format(ost_pt, 'A', a, 23, 65, 16, True)
oform(ost_pt, 'A', a, 'thick')

a += 1
b = copy.copy(a)

for i in range(3, birn.max_row):
    if re.search('задняя четвертина|передняя четвертина|ФС',
                 str(birn['B' + str(i)].value)) and birn['J' + str(i)].value != 0:
        ost_pt['A' + str(a)] = birn['B' + str(i)].value
        ost_pt['B' + str(a)] = birn['J' + str(i)].value
        a += 1

for i in range(2, ost_pt.max_row):
    try:
        ost_pt['B' + str(i)].value * 1
    except:
        ost_pt['B' + str(i)].value = 0

ost_pt['B' + str(a)] = sum([ost_pt['B' + str(i)].value for i in range(b, a) if ost_pt['B' + str(i)].value != None])

ost_pt['A' + str(a)] = 'Итого'
ost_pt['A' + str(a)].fill = PatternFill('solid', start_color='BCBCBC')
ost_pt['B' + str(a)].fill = PatternFill('solid', start_color='BCBCBC')
format(ost_pt, 'A', a, 23, 65, 16, True)
oform(ost_pt, 'A', a, 'thick')
for i in range(36, a + 1):
    oform(ost_pt, 'B', i, 'thin')
    ost_pt['B' + str(i)].font = Font(size=11, bold=True)
for i in range(1, birn.max_row):
    q = str(birn['B' + str(i)].value)
    if re.search('4 кат.* охл. без шкуры', q):

        ost_pt['B14'] = birn['J' + str(i)].value //2
    elif re.search('4 ка.*МПК',
                   q):
        ost_pt['B15'] = birn['J' + str(i)].value // 2
    elif re.search('Свинина в тушах и полутушах охл.2 кат.в шкуре',
                   q):
        ost_pt['B16'] = birn['J' + str(i)].value//2

nomer = 1
for i in range(1, birn.max_row):
    if re.search('Холодильник хранения туш', str(birn['B' + str(i)].value)):
        nomer = i

for i in range(11, nomer):

    t = '0'

    if birn['B' + str(i)].value != None:
        t = birn['B' + str(i)].value
    for j in range(3, 13):
        if ost_pt['A' + str(j)].value != None:
            s = ost_pt['A' + str(j)].value
        if s.replace('\t', '').strip() == t.replace(
                '\t', '').strip():
            ost_pt['D' + str(j)] = int(birn['J' + str(i)].value) / 2
    if re.search('ВК 1', t):
        a = np.nan if ost_pt['D29'].value is None else ost_pt['D29'].value
        ost_pt['D29'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВК 2', t):
        a = np.nan if ost_pt['D30'].value is None else ost_pt['D30'].value
        ost_pt['D30'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВК Тощ', t):
        a = np.nan if ost_pt['D31'].value is None else ost_pt['D31'].value
        ost_pt['D31'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МБК', t):
        a = np.nan if ost_pt['D19'].value is None else ost_pt['D19'].value
        ost_pt['D19'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])

    if re.search('МБ .* (Суп|Экстр|Прима)', t):
        a = np.nan if ost_pt['D23'].value is None else ost_pt['D23'].value
        ost_pt['D23'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МТ .* (Суп|Экстр|Прима)', t):
        a = np.nan if ost_pt['D25'].value is None else ost_pt['D25'].value
        ost_pt['D25'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МТ .* (Хорош|Отлич)', t):
        a = np.nan if ost_pt['D26'].value is None else ost_pt['D26'].value
        ost_pt['D26'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МБ .* (Хорош|Отлич)', t):
        a = np.nan if ost_pt['D24'].value is None else ost_pt['D24'].value
        ost_pt['D24'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МТ .* (Удовл|Низка)', t):
        a = np.nan if ost_pt['D28'].value is None else ost_pt['D28'].value
        ost_pt['D28'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МБ .* (Удовл|Низка)', t):
        a = np.nan if ost_pt['D27'].value is None else ost_pt['D27'].value
        ost_pt['D27'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВСК от МБК', t):
        a = np.nan if ost_pt['D21'].value is None else ost_pt['D21'].value
        ost_pt['D21'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВСК от МБ\b', t):
        a = np.nan if ost_pt['D20'].value is None else ost_pt['D20'].value
        ost_pt['D20'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВСК от МТ', t):
        a = np.nan if ost_pt['D22'].value is None else ost_pt['D22'].value
        ost_pt['D22'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
for i in range(nomer, birn.max_row):

    t = '0'

    if birn['B' + str(i)].value != None:
        t = birn['B' + str(i)].value
    if re.search('ВК 1.* в полуту', t):
        a = np.nan if ost_pt['B29'].value is None else ost_pt['B29'].value
        ost_pt['B29'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВК 2.* в полуту', t):
        a = np.nan if ost_pt['B30'].value is None else ost_pt['B30'].value
        ost_pt['B30'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВК Тощ.* в полуту', t):
        a = np.nan if ost_pt['B31'].value is None else ost_pt['B31'].value
        ost_pt['B31'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МБК', t) and not re.search('ВСК', t):
        a = np.nan if ost_pt['B19'].value is None else ost_pt['B19'].value
        ost_pt['B19'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])

    if re.search('МБ .* (Суп|Экстр|Прима)', t):
        a = np.nan if ost_pt['B23'].value is None else ost_pt['B23'].value
        ost_pt['B23'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МТ .* (Суп|Экстр|Прима)', t):
        a = np.nan if ost_pt['B25'].value is None else ost_pt['B25'].value
        ost_pt['B25'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МТ .* (Хорош|Отлич)', t):
        a = np.nan if ost_pt['B26'].value is None else ost_pt['B26'].value
        ost_pt['B26'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МБ .* (Хорош|Отлич)', t):
        a = np.nan if ost_pt['B24'].value is None else ost_pt['B24'].value
        ost_pt['B24'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МТ .* (Удовл|Низка)', t):
        a = np.nan if ost_pt['B28'].value is None else ost_pt['B28'].value
        ost_pt['B28'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('МБ .* (Удовл|Низка)', t):
        a = np.nan if ost_pt['B27'].value is None else ost_pt['B27'].value
        ost_pt['B27'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВСК от МБК', t):
        a = np.nan if ost_pt['B21'].value is None else ost_pt['B21'].value
        ost_pt['B21'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
    if re.search('ВСК от МБ кат', t):
        a = np.nan if ost_pt['B20'].value is None else ost_pt['B20'].value
        ost_pt['B20'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])

    if re.search('ВСК от МТ', t):
        a = np.nan if ost_pt['B22'].value is None else ost_pt['B22'].value
        ost_pt['B22'] = np.nansum([a, int(birn['J' + str(i)].value) / 2])
ost_pt['B32'] = sum([ost_pt['B' + str(i)].value for i in range(19,32)])
ost_pt['B32'].font = Font(size=12,bold=True)
kat1 = '0'
for i in range(nomer, birn.max_row + 1):
    if re.search('беконная', str(birn['B' + str(i)].value)):
        kat1 = i

    t = '0'
    if birn['B' + str(i)].value != None:
        t = birn['B' + str(i)].value
    for j in range(3, 13):
        if ost_pt['A' + str(j)].value != None:
            s = ost_pt['A' + str(j)].value
        if s.replace('\t', '').strip() == t.replace(
                '\t', '').strip():
            ost_pt['B' + str(j)] = int(birn['J' + str(i)].value) / 2

ost_pt['B7'] = sum((int(ost_pt['B' + str(i)].value) for i in range(3, 7)))
ost_pt['B10'] = sum((int(ost_pt['B' + str(i)].value) for i in range(8, 10)))
ost_pt['B13'] = sum((int(ost_pt['B' + str(i)].value) for i in [7, 10, 11, 12]))
s = []
if kat1 != '0':
    q = birn['B' + str(kat1 + 1)].value
    try:
        while q == None  or re.search('\d{2}.\d{2}.\d{4}', q):
            if birn['J' + str(kat1 + 1)].value != 0:
                s.append([birn['B' + str(kat1 + 1)].value, birn['J' + str(kat1 + 1)].value / 2])
            kat1 += 1
            q = birn['B' + str(kat1 + 1)].value
    except:
        print(q,kat1)


for i in range(len(s)):
    if s[i][0] != None:
        dt = datetime.datetime.strptime(s[i][0], '%d.%m.%Y %H:%M:%S')
        dt -= datetime.timedelta(days=1)
        s[i][0] = dt.strftime('%d.%m.%Y')

s = str([f'{s[i][0]} - {s[i][1]}' for i in range(len(s))])
ost_pt['C11'] = s
for i in range(3, 13):
    ost_pt['A' + str(i)] = list(sv.values())[i - 3]
ost_pt['A6'] = '2 кат дюрки'


def clicked():
    ost_pt['B6'] = int(count.get())
    window.destroy()


window = Tk()
window.title("Количество ТУШ 2 категории дюрков?")
window.geometry('600x600')
lbl = Label(window, text="Количество ТУШ 2 категории дюрков?", font=("Arial Bold", 15))
lbl.pack()

count = Entry(window)
count.pack()
btn = Button(window, text="ОК!", command=clicked).pack()

window.mainloop()

ost_pt['B3'] = ost_pt['B3'].value - ost_pt['B6'].value
for i in range(nomer, birn.max_row):
    q = birn['B' + str(i)].value
    if q != None and re.search('\d{2}.\d{2}.\d{4}', q):
        dt = datetime.datetime.strptime(q, '%d.%m.%Y %H:%M:%S')
        try:
            dt = dt - datetime.timedelta(days=1)
        except OverflowError:
            birn['B' + str(i)] = '0'

        birn['B' + str(i)] = dt.strftime('%d.%m.%Y')

for i in range(19, 32):
    ost_pt['D' + str(i)].fill = PatternFill('solid', start_color="F7FF00")
    format(ost_pt, 'D', i, 18, 36, 11, True)
    oform(ost_pt, 'D', i, 'thin')

vk = {}
repile = 'Говядина от кат. охл. в полутушах с вырезкой'.split()
for i in range(nomer, birn.max_row):
    vkDate = []
    vkCount = []
    q = str(birn['B' + str(i)].value)

    if re.search('ВК .* в полутушах', q):

        ind = i + 1
        while birn['B' + str(ind)].value == None or re.search('\d{2}.\d{2}.\d{4}', str(birn['B' + str(ind)].value)):
            if birn['J' + str(ind)].value != 0:
                vkDate.append(birn['B' + str(ind)].value)
                vkCount.append(int(birn['J' + str(ind)].value) / 2)
            ind += 1
            d = []
        for i in q.split():  # Обрезание названия

            if not i in repile:
                d.append(i)
            q = ''.join(d)
        vk[q] = pd.Series(vkCount, index=vkDate)
vkd = pd.DataFrame(vk)
vkd.index = pd.to_datetime(vkd.index, errors='ignore',dayfirst=True)
vkd = vkd.sort_index()
vkd.index = vkd.index.strftime('%d.%m.%Y')


pam.append('Распредели в остатках кости - перетяни в колонку заморозка, если в клетках ')
pam.append('Распредели ливер, если есть в заморозке ')
pam.append('Проверь, нет ли в твоих остатках отрицательных значений. ')

datesk = list(vkd.index)

nomsk = list(vkd.columns)

entrylist = []
total_rows = len(datesk)
total_columns = len(nomsk)

root = Tk()
root.title('Сан.Брак коровы')

for i in range(total_rows):
    lbl = Label(root, text=datesk[i], width=20, relief=RIDGE, fg='blue', font=('Arial', 16, 'bold')).grid(row=i + 1,
                                                                                                          column=0)

vks = {}
for i in range(total_columns):
    vksDate = []
    vksCount = []
    lbl = Label(root, text=nomsk[i], anchor='w', width=20, justify=CENTER, relief=RIDGE, fg='blue',
                font=('Arial', 16, 'bold'))
    lbl.grid(row=0, column=i + 1)
    for j in range(total_rows):
        entry = Entry(root, width=20, fg='blue', font=('Arial', 16, 'bold'))
        entry.grid(row=j + 1, column=i + 1)
        vksCount.append(entry)
        vksDate.append(datesk[j])

        entry.insert(END, 0)
    vks[nomsk[i]] = pd.Series(vksCount, index=vksDate)
vks = pd.DataFrame(vks)
vks.index = pd.to_datetime(vks.index, errors='ignore',dayfirst=True)
vks = vks.sort_index()
vks.index = vks.index.strftime('%d.%m.%Y')

def click_btn():
    for col in list(vks.columns):
        for date in list(vks.index):
            vks.at[date, col] = int(vks.loc[date, col].get())
    root.destroy()


btn = Button(root, text="Отправить!", command=click_btn, width=20, bg='#6cd9ca',
             font=('Novartis Deco', 16, 'bold')).grid(row=0, column=0)

root.mainloop()
vkd = vkd.replace(np.nan, 0)
vkd -= vks
vkd = vkd.replace(0, np.nan)




if [j for i in vkd.values for j in i if j < 0]:
    otrKorov = Tk()
    otrKorov.geometry('600x600')
    lbl = Label(otrKorov, text='Неверные данные по коровам: Сан.Брак верно указан?').pack()
    button = Button(otrKorov, text="Принял", command=otrKorov.destroy).pack()

    otrKorov.mainloop()

vksK = [('Сан.брак', str(i)) for i in vks.columns]
vks.columns = pd.MultiIndex.from_tuples(vksK)
vkd.loc[''] = np.nan
vkd.loc['ИТОГО'] = vkd.sum()

vks.loc[''] = np.nan
vks.loc['ИТОГО'] = vks.sum()
vks = vks.replace(0, np.nan)
ubsv = 0
ubvsk = 0
ubmbk = 0
ubmb = 0
ubmt = 0
ubvk = 0
for i in range(1, birn.max_row):
    q = str(birn['B' + str(i)].value)
    qq = birn['H' + str(i)].value
    if re.search('баками', q) and qq != None:
        ubsv += int(qq) / 2
    elif re.search('ВСК', q) and qq != None:
        ubvsk += int(qq) / 2
    elif re.search('МБК', q) and qq != None:
        ubmbk += int(qq) / 2
    elif re.search('МБ', q) and qq != None:
        ubmb += int(qq) / 2
    elif re.search('МТ.*в полутушах', q) and qq != None:
        ubmt += int(qq) / 2
    elif re.search('ВК.*в полутушах', q) and qq != None:
        ubvk += int(qq) / 2
rcv = 0
vksk = 0
svsk = 0
for i in range(1, nomer):
    q = str(birn['B' + str(i)].value)
    qq = birn['I' + str(i)].value
    if re.search('ВК.*в полутушах', q) and qq != None:
        vksk += int(qq) / 2
    if re.search('баками', q) and qq != None:
        svsk += int(qq) / 2
ubsv -= svsk
ubvk -= vksk
for i in range(nomer, birn.max_row):
    q = str(birn['B' + str(i)].value)
    qq = birn['I' + str(i)].value
    if re.search('баками', q) and qq != None:
        rcv += int(qq) / 2


pam.append(f'Расход по всем участкам составил {rcv} свиней')
if ubvk != 0:
    pam.append(f' Убой Коров составил {ubvk} голов. ')
if ubmt != 0:
    pam.append(f' Убой Телок составил {ubmt} голов. ')
if ubmb != 0:
    pam.append(f' Убой Быков составил {ubmb} голов. ')
if ubvsk != 0:
    pam.append(f' Убой ВСК составил {ubvsk} голов. ')
if ubmbk != 0:
    pam.append(f' Убой кастратов составил {ubmbk} голов. ')
pam.append(
    f' Убой составил {ubsv} свиней,это примерно {round(ubsv / 160)} машин, Возможная погрешность - приход с сан.камеры. ')
for i in pam:
    memory.append([i])

ost.save(f'{name}.xlsx')

properties = {"border": "2px solid black", "font-size": "28px", 'width': '100'}

with pd.ExcelWriter(f'{name}.xlsx', mode="a", if_sheet_exists='replace', engine="openpyxl") as writer:
    vkd.style.set_properties(**properties).to_excel(writer, sheet_name="Остаток по датам коровы", startcol=1,
                                                    startrow=1)

with ExcelWriter(f'{name}.xlsx',
                 mode="a",
                 engine="openpyxl",
                 if_sheet_exists="overlay",
                 ) as writer:
    vks.style.set_properties(**properties).to_excel(writer, sheet_name="Остаток по датам коровы",
                                                    startrow=len(vkd.index) + 5, startcol=1)
mb = {}
repile = 'Говядина от кат. охл. в полутушах с вырезкой'.split()
molod = []
for i in range(nomer, birn.max_row):
    mbDate = []
    mbCount = []
    q = str(birn['B' + str(i)].value)
    qq = birn['J' + str(i)].value
    if re.search('МБ\s.* в полутушах', q) and not re.search('ВСК', q) and qq != 0:

        ind = i + 1
        while birn['B' + str(ind)].value == None or re.search('\d{2}.\d{2}.\d{4}', str(birn['B' + str(ind)].value)):
            if birn['J' + str(ind)].value != 0:
                mbDate.append(birn['B' + str(ind)].value)
                mbCount.append(int(birn['J' + str(ind)].value) / 2)
            ind += 1
            d = []
        for i in q.split():  # Обрезание названия

            if not i in repile:
                d.append(i)
            q = ' '.join(d)
        mb[q] = pd.Series(mbCount, index=mbDate)
mb = pd.DataFrame(mb)
mb.index = pd.to_datetime(mb.index, errors='ignore',dayfirst=True)
mb = mb.sort_index()
mb.index = mb.index.strftime('%d.%m.%Y')
mbs = ['МБ Супер', 'МБ Прима', 'МБ Экстра', 'МБ Отличная', 'МБ Хорошая', 'МБ Удовлетворительная', 'МБ Низкая']

mb = mb.reindex(columns=mbs)
mbK = [('Первая категория', i) for i in mb.columns[:3]] + [('Вторая категория', i) for i in mb.columns[3:5]] + [
    ('Тощая', i) for i in mb.columns[5:]]
mb = mb.rename_axis('Дата')
mb.columns = pd.MultiIndex.from_tuples(mbK)
mb.loc[''] = np.nan

mb.loc['ИТОГО'] = mb.sum()
molod.append(mb)
mt = {}
for i in range(nomer, birn.max_row):
    mtDate = []
    mtCount = []
    q = str(birn['B' + str(i)].value)
    qq = birn['J' + str(i)].value
    if re.search('МТ\s.* в полутушах', q) and not re.search('ВСК', q) and qq != 0:

        ind = i + 1
        while birn['B' + str(ind)].value == None or re.search('\d{2}.\d{2}.\d{4}', str(birn['B' + str(ind)].value)):
            if birn['J' + str(ind)].value != 0:
                mtDate.append(birn['B' + str(ind)].value)
                mtCount.append(int(birn['J' + str(ind)].value) / 2)
            ind += 1
            d = []
        for i in q.split():  # Обрезание названия

            if not i in repile:
                d.append(i)
            q = ' '.join(d)
        mt[q] = pd.Series(mtCount, index=mtDate)
mt = pd.DataFrame(mt)
mt.index = pd.to_datetime(mt.index, errors='ignore',dayfirst=True)
mt = mt.sort_index()
mt.index = mt.index.strftime('%d.%m.%Y')
mts = ['МТ Супер', 'МТ Прима', 'МТ Экстра', 'МТ Отличная', 'МТ Хорошая', 'МТ Удовлетворительная', 'МТ Низкая']

mt = mt.reindex(columns=mts)
mtK = [('Первая категория', i) for i in mt.columns[:3]] + [('Вторая категория', i) for i in mt.columns[3:5]] + [
    ('Тощая', i) for i in mt.columns[5:]]

mt.columns = pd.MultiIndex.from_tuples(mtK)
mt = mt.rename_axis('Дата')
molod.append(mt)
mbk = {}
for i in range(nomer, birn.max_row):
    mtDate = []
    mtCount = []
    q = str(birn['B' + str(i)].value)
    qq = birn['J' + str(i)].value
    if re.search('МБК\s.* в полутушах', q) and not re.search('ВСК', q) and qq != 0:

        ind = i + 1
        while birn['B' + str(ind)].value == None or re.search('\d{2}.\d{2}.\d{4}', str(birn['B' + str(ind)].value)):
            if birn['J' + str(ind)].value != 0:
                mtDate.append(birn['B' + str(ind)].value)
                mtCount.append(int(birn['J' + str(ind)].value) / 2)
            ind += 1
            d = []
        for i in q.split():  # Обрезание названия

            if not i in repile:
                d.append(i)
            q = ' '.join(d)
        mbk[q] = pd.Series(mtCount, index=mtDate)
mbk = pd.DataFrame(mbk)
mbk.index = pd.to_datetime(mbk.index, errors='ignore',dayfirst=True)
mbk = mbk.sort_index()
mbk.index = mbk.index.strftime('%d.%m.%Y')

mbks = ['МБКСупер', 'МБК Прима', 'МБК Экстра', 'МБК Отличная', 'МБК Хорошая', 'МБК Удовлетворительная', 'МБК Низкая']
mbk = mbk.rename_axis('Дата')
mbk = mbk.reindex(columns=mbks)
mbkK = [('Первая категория', i) for i in mbk.columns[:3]] + [('Вторая категория', i) for i in mbk.columns[3:5]] + [
    ('Тощая', i) for i in mbk.columns[5:]]

mbk.columns = pd.MultiIndex.from_tuples(mbkK)
molod.append(mbk)
mtvsk = {}
for i in range(nomer, birn.max_row):
    mtDate = []
    mtCount = []
    q = str(birn['B' + str(i)].value)
    qq = birn['J' + str(i)].value
    if re.search('МТ\s.* в полутушах', q) and re.search('ВСК', q) and qq != None:

        ind = i + 1
        while birn['B' + str(ind)].value == None or re.search('\d{2}.\d{2}.\d{4}', str(birn['B' + str(ind)].value)):
            if birn['J' + str(ind)].value != 0:
                mtDate.append(birn['B' + str(ind)].value)
                mtCount.append(int(birn['J' + str(ind)].value) / 2)
            ind += 1
            d = []
        for i in q.split():  # Обрезание названия

            if not i in repile:
                d.append(i)
            q = ' '.join(d)
        mtvsk[q] = pd.Series(mtCount, index=mtDate)
mtvsk = pd.DataFrame(mtvsk)
mtvsk.index = pd.to_datetime(mtvsk.index, errors='ignore',dayfirst=True)
mtvsk = mtvsk.sort_index()
mtvsk.index = mtvsk.index.strftime('%d.%m.%Y')
mtvsk = mtvsk.rename_axis('Дата')
molod.append(mtvsk)
mbkvsk = {}
for i in range(nomer, birn.max_row):
    mtDate = []
    mtCount = []
    q = str(birn['B' + str(i)].value)
    qq = birn['J' + str(i)].value
    if re.search('МБК\s.* в полутушах', q) and re.search('ВСК', q) and qq != 0:

        ind = i + 1
        while birn['B' + str(ind)].value == None or re.search('\d{2}.\d{2}.\d{4}', str(birn['B' + str(ind)].value)):
            if birn['J' + str(ind)].value != 0:
                mtDate.append(birn['B' + str(ind)].value)
                mtCount.append(int(birn['J' + str(ind)].value) / 2)
            ind += 1
            d = []
        for i in q.split():  # Обрезание названия

            if not i in repile:
                d.append(i)
            q = ' '.join(d)
        mbkvsk[q] = pd.Series(mtCount, index=mtDate)
mbkvsk = pd.DataFrame(mbkvsk)
mbkvsk.index = pd.to_datetime(mbkvsk.index, errors='ignore',dayfirst=True)
mbkvsk = mbkvsk.sort_index()
mbkvsk.index = mbkvsk.index.strftime('%d.%m.%Y')
mbkvsk = mbkvsk.rename_axis('Дата')
molod.append(mbkvsk)
mbvsk = {}
for i in range(nomer, birn.max_row):
    mtDate = []
    mtCount = []
    q = str(birn['B' + str(i)].value)
    qq = birn['J' + str(i)].value
    if re.search('ВСК от МБ\s.* в полутушах', q) and re.search('ВСК', q) and qq != 0:

        ind = i + 1
        while birn['B' + str(ind)].value == None or re.search('\d{2}.\d{2}.\d{4}', str(birn['B' + str(ind)].value)):
            if birn['J' + str(ind)].value != 0:
                mtDate.append(birn['B' + str(ind)].value)
                mtCount.append(int(birn['J' + str(ind)].value) / 2)
            ind += 1
            d = []
        for i in q.split():  # Обрезание названия

            if not i in repile:
                d.append(i)
            q = ' '.join(d)
        mbvsk[q] = pd.Series(mtCount, index=mtDate)
mbvsk = pd.DataFrame(mbvsk)
mbvsk.index = pd.to_datetime(mbvsk.index, errors='ignore',dayfirst=True)
mbvsk = mbvsk.sort_index()
mbvsk.index = mbvsk.index.strftime('%d.%m.%Y')
mbvsk = mbvsk.rename_axis('Дата')
#print(sum(mbvsk))
molod.append(mbvsk)

ind = len(mb.index) + 6
dtmolod = [ind]
mb = mb.replace(0, np.nan)

with pd.ExcelWriter(f'{name}.xlsx', mode="a", if_sheet_exists='replace', engine="openpyxl") as writer:
    mb.style.set_properties(**properties).to_excel(writer, sheet_name="Остатки по датам молодняк", startcol=1,
                                                   startrow=1)
for i in molod[1:]:
    if not i.empty:
        i.loc[''] = np.nan
        i.loc['Итого'] = i.sum()
        with ExcelWriter(f'{name}.xlsx', mode="a", engine="openpyxl", if_sheet_exists="overlay", ) as writer:
            i = i.replace(0, np.nan)
            i.style.set_properties(**properties).to_excel(writer, sheet_name="Остатки по датам молодняк", startrow=ind,
                                                          startcol=1)
            ind += len(i.index) + 5
            dtmolod.append(ind)

ost = load_workbook(f'{name}.xlsx')
ost_vk = ost["Остаток по датам коровы"]
ost_vk['B2'] = 'Дата'
ost_vk['B' + str(len(vkd.index) + 7)] = "Дата"
ost_vk.delete_rows(len(vkd.index) + 8)
ost_vk.column_dimensions['B'].width = 19
ost_vk.column_dimensions['C'].width = 13
ost_vk.column_dimensions['D'].width = 13
ost_vk.column_dimensions['E'].width = 25
ost_vk.column_dimensions['F'].width = 47
for i in range(1, ost_vk.max_row + 1):
    ost_vk.row_dimensions[i].height = 34
    if ost_vk['B' + str(i)].value:
        ost_vk['B' + str(i)].fill = PatternFill('solid', start_color="F0FF33")

    for j in range(1, ost_vk.max_column + 1):
        ost_vk.cell(i, j).font = Font(size=16, bold=True)

ost_mb = ost["Остатки по датам молодняк"]

for i in range(1, ost_mb.max_row):
    if ost_mb['B' + str(i)].value == 'Дата' and ost_mb['C' + str(i)].value == None:
        ost_mb['B' + str(i - 1)] = 'Дата'
        s = 'B' + str(i - 1)
        ss = 'B' + str(i)
        ost_mb.merge_cells(f'{s}:{ss}')

ost_mb.column_dimensions['B'].width = 20
ost_mb.column_dimensions['C'].width = 20
ost_mb.column_dimensions['D'].width = 20
ost_mb.column_dimensions['E'].width = 20
ost_mb.column_dimensions['F'].width = 25
ost_mb.column_dimensions['G'].width = 24
ost_mb.column_dimensions['H'].width = 45
ost_mb.column_dimensions['I'].width = 20
for i in range(1, ost_mb.max_row + 1):
    ost_mb.row_dimensions[i].height = 34
    if ost_mb['B' + str(i)].value:
        ost_mb['B' + str(i)].fill = PatternFill('solid', start_color="F0FF33")

    for j in range(1, ost_mb.max_column + 1):
        ost_mb.cell(i, j).font = Font(size=16, bold=True)

ost.save(f'{name}.xlsx')

